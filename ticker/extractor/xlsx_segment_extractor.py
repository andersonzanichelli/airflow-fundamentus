from pandas import DataFrame
from openpyxl import load_workbook

from ticker.extractor.segment_extractor import SegmentExtractor
from ticker.model.segment import SegmentBuilder

class XlsxSegmentExtractor(SegmentExtractor):

    def __init__(self, sheet) -> None:
        super().__init__()
        self.sheet = sheet
    
    def extract(self, filepath: str) -> DataFrame:

        wb = load_workbook(filename=filepath, read_only=True)
        ws = wb[self.sheet]
        segments = self.build(ws)
        wb.close()

        df = DataFrame.from_dict(segments)
        print(df.head())
        return df

    def build(self, ws):
        max = ws.max_row
        row = 1
        segments = []

        while(row < max):

            if(ws[f"A{row}"].value is not None):
                sector = ws[f"A{row}"].value

            if(ws[f"B{row}"].value is not None):
                subsector = ws[f"B{row}"].value

            if(ws[f"C{row}"].value is not None and ws[f"D{row}"].value is None):
                segment = ws[f"C{row}"].value
            
            row = row + 1

            while(ws[f"A{row}"].value is None and ws[f"B{row}"].value is None and ws[f"C{row}"].value is not None and ws[f"D{row}"].value is not None):
                builder = SegmentBuilder() \
                                    .set_sector(sector) \
                                    .set_subsector(subsector) \
                                    .set_segment(segment) \
                                    .set_code(ws[f"D{row}"].value) \
                                    .set_segcode(ws[f"E{row}"].value)
                
                stock = builder.build()
                segments.append(stock.asdict())
                row = row + 1
        
        return segments